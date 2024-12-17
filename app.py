import os
import time
from datetime import datetime
from flask import Flask, render_template, request, session, redirect, url_for
from openpyxl import Workbook, load_workbook
from extraer_texto import extraer_texto_pdf
from procesar_texto import procesar_texto
from generar_embeddings import cargar_modelo_embeddings, generar_embeddings
from buscar_contexto import buscar_oraciones_similares
from generar_respuesta import configurar_openai, generar_respuesta
import numpy as np

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "clave_por_defecto")

# Configuración de costes por token de OpenAI
COSTO_POR_1K_TOKENS = 0.002  # Coste estimado en USD para GPT-3.5-turbo

# Configurar la API de OpenAI
configurar_openai()

# Cargar modelo y embeddings al iniciar el servidor
modelo = cargar_modelo_embeddings()
if os.path.exists('embeddings.npy') and os.path.exists('oraciones.npy'):
    embeddings_oraciones = np.load('embeddings.npy')
    oraciones = np.load('oraciones.npy', allow_pickle=True)
else:
    ruta_pdf = 'asignatura.pdf'
    texto_pdf = extraer_texto_pdf(ruta_pdf)
    oraciones = procesar_texto(texto_pdf)
    embeddings_oraciones = generar_embeddings(oraciones, modelo)
    np.save('embeddings.npy', embeddings_oraciones)
    np.save('oraciones.npy', oraciones)

# Archivo para almacenar métricas
METRICS_FILE = "chatbot_metrics.xlsx"

# Inicializar el archivo de métricas con encabezados si no existe
if not os.path.exists(METRICS_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Métricas"
    ws.append([
        "Timestamp", "Pregunta", "Respuesta", "Rol", "Idioma", "Tiempo de Respuesta (s)",
        "Longitud Pregunta", "Longitud Respuesta", "Tokens Prompt", 
        "Tokens Completion", "Tokens Total", "Coste (USD)", "Error"
    ])
    wb.save(METRICS_FILE)

@app.route('/set_language/<language>')
def set_language(language):
    session['idioma'] = language
    return redirect(url_for('index'))

@app.route('/', methods=['GET', 'POST'])
def index():
    current_year = datetime.now().year
    idioma = session.get('idioma', 'es')
    if 'historial' not in session:
        session['historial'] = []

    if request.method == 'POST':
        pregunta = request.form['pregunta']
        rol = request.form['rol']
        inicio_procesamiento = time.time()  # Inicio del cálculo de tiempo
        respuesta = ""
        error = None
        tokens_prompt = 0
        tokens_completion = 0
        tokens_total = 0
        coste = 0.0

        try:
            # Buscar contexto y generar respuesta
            contexto = buscar_oraciones_similares(pregunta, embeddings_oraciones, oraciones, modelo)
            respuesta_objeto = generar_respuesta(contexto, pregunta, rol, idioma)
            respuesta = respuesta_objeto["respuesta"]
            tokens_prompt = respuesta_objeto["tokens_prompt"]
            tokens_completion = respuesta_objeto["tokens_completion"]
            tokens_total = respuesta_objeto["tokens_total"]
            coste = respuesta_objeto["coste"]
        except Exception as e:
            respuesta = "Lo siento, ocurrió un error al procesar tu solicitud." if idioma == 'es' else "Sorry, an error occurred while processing your request."
            error = str(e)

        # Calcular métricas
        tiempo_respuesta = time.time() - inicio_procesamiento
        longitud_pregunta = len(pregunta)
        longitud_respuesta = len(respuesta)

        # Guardar métricas en Excel
        if os.path.exists(METRICS_FILE):
            wb = load_workbook(METRICS_FILE)
            ws = wb["Métricas"]
            ws.append([
                datetime.now().isoformat(),
                pregunta,
                respuesta,
                rol,
                idioma,
                round(tiempo_respuesta, 2),
                longitud_pregunta,
                longitud_respuesta,
                tokens_prompt,
                tokens_completion,
                tokens_total,
                round(coste, 4),
                error
            ])
            wb.save(METRICS_FILE)

        # Agregar al historial
        session['historial'].append({
            'pregunta': pregunta,
            'respuesta': respuesta,
            'rol': rol.capitalize()
        })
        session.modified = True

    else:
        rol = ''
    return render_template('index.html', historial=session['historial'], rol=rol, idioma=idioma, current_year=current_year)

if __name__ == '__main__':
    app.run(debug=True)
